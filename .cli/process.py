import re
from openpyxl import load_workbook
from pprint import pprint
from transliterate import translit

def parse_xls(file_path, sheet_name=None):
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    result = []
    current_category = None

    for row in ws.iter_rows():
        values = [cell.value for cell in row]

        # Прерывание при полностью пустой строке
        if all(val is None or str(val).strip() == '' for val in values):
            print('⏹️ Пустая строка достигнута — остановка')
            break

        # Категория (данные в первых ячейках)
        if any(values[0:4]):
            category = {
                'category_data': values[0:4],
                'subcategories': []
            }
            result.append(category)
            current_category = category
        # Подкатегория (данные с 5-й ячейки)
        elif any(values[4:]) and current_category:
            current_category['subcategories'].append(values[4:])

    return result

def slugify(text):
    # Транслитерация (ru → latin)
    text = translit(text, 'ru', reversed=True)

    # Приведение к нижнему регистру
    text = text.lower()

    # Замена пробелов и любых не-букв/цифр на тире
    text = re.sub(r'\s+', '-', text)           # пробелы → тире
    text = re.sub(r'[^a-z0-9\-]', '', text)     # удаление лишнего
    text = re.sub(r'-{2,}', '-', text)          # двойные тире → одно
    text = text.strip('-')                      # обрезка тире по краям

    return "-" + text

parsed = parse_xls('d4_2.xlsx')
descriptions = [
    (
        'В нашем интернет-магазине вы можете купить запчасти на {d_br_model} по доступным ценам с доставкой по всей территории РБ, РФ и Казахстана. '
        'Заказы доставляются собственным транспортном, почтой и профильными транспортными компаниями. С подробной информацией по этому поводу можно '
        'ознакомиться <a href="https://d4.by/delivery">тут</a>. Оплата производится по наличному и безналичному расчёту. На каждую запчасть '
        'предоставляется индивидуальная гарантия, с деталями которой вы можете ознакомиться <a href="https://d4.by/garantyja">здесь</a>.'
    ),
    (
        'Интернет-магазин D4.by предлагает качественные бу запчасти {d_br_model} по выгодным ценам с гарантией и доставкой по Беларуси, России и '
        'Казахстану. Гарантия на каждый товар предоставляется индивидуально. Ознакомится с её условиями вы можете по <a href="https://d4.by/garantyja">этой ссылке'
        '</a>. Заказы доставляются собственным транспортом, почтой и профильными транспортными компаниями. Произвести оплату можно через наличный и безналичный '
        'расчёт. Подробная информация о доставке и оплате <a href="https://d4.by/delivery">тут</a>'
    ),
    (
        'В интернет-магазине D4.by можно купить бу запчасти {d_br_model} из Европы по демократичным ценам. Заказы доставляются по всей Беларуси, России и '
        'Казахстану. Оплатить покупку можно наличным и безналичным расчётом. Подробнее о доставке и оплате <a href="https://d4.by/delivery">тут</a>. На '
        'каждый товар распространяется индивидуальная гарантия, с условиями которой можно ознакомиться по <a href="https://d4.by/garantyja">ссылке</a>.'
    )
]
var_id = -1;

for category in parsed:
    cat_values = category['category_data']

    for sub in category['subcategories']:
        var_id = var_id + 1 if var_id < 2 else 0
        description = descriptions[var_id]
        alias = 'zapchasti-' + cat_values[3] + slugify(str(sub[0]))
        mfp = f'c-kategorii-0[{cat_values[1]},{sub[1]}]'
        h1 = f'Запчасти {cat_values[2]} ({cat_values[0]}) '
        d_br_model = f'{cat_values[2]} ({cat_values[0]}) '
        meta_description = f'Запчасти {cat_values[2]} ({cat_values[0]}) '
        meta_keywords = f'запчасти {cat_values[2]} {sub[0]}, запчасти {cat_values[2]} {sub[0]} цены, запчасти {cat_values[2]} {sub[0]} купить'
        # если русского языка нет, берем английский
        if (sub[2] is None or str(sub[2]).strip() == ''):
            d_br_model += str(sub[0])
            h1 += str(sub[0])
            meta_description += str(sub[0])
        else:
           d_br_model += str(sub[2]) + f' ({str(sub[0])})'
           h1 += str(sub[2]) + f' ({str(sub[0])})'
           meta_description += str(sub[2]) + f' ({str(sub[0])})'

        meta_title = h1 + ' купить цены'
        meta_description += ' купить в интернет-магазине с доставкой по доступным ценам: ☎ +375 (29) 800-07-36, +375 (29) 800-07-36.'
        description = description.format(d_br_model=d_br_model)
        
        sql = (
            'insert into `oc_mfilter_url_alias` (path, mfp, alias, language_id, store_id, meta_title, meta_description, meta_keyword, description, h1) '
            f'values (\'katalog\', \'{mfp}\',\'{alias}\', 1, 0, \'{meta_title}\', \'{meta_description}\', \'{meta_keywords}\', \'{description}\', \'{h1}\' );'
        )

        print(sql)
